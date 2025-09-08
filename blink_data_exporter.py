import logging
import os
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from frame_info import Eye, FrameInfo, EyeData


class BlinkDataExporter:
    def __init__(
        self,
        session_save_dir: str,
        frame_rate: float,
        eyes: Optional[List[Eye]] = None,
        ground_truth_df: Optional[pd.DataFrame] = None,
        video_id: int = 1,
    ):
        logging.info("Initializing BlinkDataExporter.")
        self.session_save_dir = session_save_dir
        self.frame_rate = frame_rate
        self.eyes: List[Eye] = eyes if eyes is not None else [Eye.LEFT, Eye.RIGHT]
        self.ground_truth_df = ground_truth_df
        self.video_id = video_id

    def export_all_blink_data_to_excel(
        self, processed_frames: List[FrameInfo]
    ) -> pd.DataFrame:
        logging.info("Exporting all blink data to Excel.")

        excel_file_path = os.path.join(self.session_save_dir, "blink_data.xlsx")
        logging.info(f"Excel file path: {excel_file_path}")

        frame_data = self._generate_frame_data(processed_frames)

        with pd.ExcelWriter(excel_file_path, engine="openpyxl") as writer:
            frame_data.to_excel(writer, sheet_name="Frame Predictions", index=False)

        self._adjust_column_widths(excel_file_path)
        logging.info("Finished exporting all blink data to Excel.")
        return frame_data

    def generate_report_from_csv(self, csv_file_path: str) -> pd.DataFrame:
        logging.info(f"Generating report from CSV: {csv_file_path}")
        processed_frames = self.read_csv_and_convert_to_frameinfo(csv_file_path)
        return self.export_all_blink_data_to_excel(processed_frames)

    def read_csv_and_convert_to_frameinfo(self, csv_file_path: str) -> List[FrameInfo]:
        logging.info(f"Reading CSV and converting to FrameInfo: {csv_file_path}")
        df = pd.read_csv(csv_file_path, sep=";")

        frame_info_list: List[FrameInfo] = []
        for _, row in df.iterrows():
            eyes_data: dict[Eye, EyeData] = {}
            for eye in self.eyes:
                eye_str = eye.value.capitalize()
                pred_col = f"{eye_str} Eye Blink Prediction"
                blink_prob_col = f"{eye_str} Eye Blink Probability"
                closed_prob_col = f"{eye_str} Eye Closed Probability"

                if pred_col in df.columns:
                    eyes_data[eye] = EyeData(
                        img=None,  # We don't have the image here
                        pred=row.get(pred_col),
                        blink_prob=row.get(blink_prob_col),
                        closed_prob=row.get(closed_prob_col),
                    )

            frame_info = FrameInfo(
                frame_num=row["Frame Number"],
                frame_img=None,
                frame_with_boxes=None,
                eye_boxes=[],
                eyes=eyes_data,
            )
            frame_info_list.append(frame_info)

        return frame_info_list

    @staticmethod
    def _adjust_column_widths(excel_file_path):
        logging.debug(f"Adjusting column widths for {excel_file_path}")
        workbook = load_workbook(excel_file_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for column in sheet.columns:
                max_length = 0
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = max_length + 2  # Adding a little extra space
                sheet.column_dimensions[
                    get_column_letter(column[0].column)
                ].width = adjusted_width

        workbook.save(excel_file_path)

    def _generate_frame_data(self, processed_frames: List[FrameInfo]) -> pd.DataFrame:
        """Generates a DataFrame with predictions and ground truth if available."""
        records = []
        for frame in processed_frames:
            for eye in self.eyes:
                eye_data = frame.eyes.get(eye)
                record = {
                    "video": self.video_id,
                    "frameId": frame.frame_num,
                    "eye": eye.value.upper(),
                    "pred_blink": eye_data.pred if eye_data else None,
                    "pred_blink_prob": eye_data.blink_prob if eye_data else None,
                    "pred_closed_prob": eye_data.closed_prob if eye_data else None,
                }
                records.append(record)

        df = pd.DataFrame(records)

        # Generate blink IDs for predictions
        df = self.generate_blink_ids(df, "pred_blink", "pred_blink_id")

        if self.ground_truth_df is not None:
            # Filter ground truth for the specific eye(s) being processed
            eye_strs = [eye.value.upper() for eye in self.eyes]
            gt_df = self.ground_truth_df[
                self.ground_truth_df["eye"].isin(eye_strs)
            ].copy()

            df = pd.merge(
                df,
                gt_df,
                on=["frameId", "eye"],
                how="left",
                suffixes=("", "_gt"),
            )
            df.rename(
                columns={
                    "blink": "gt_blink",
                    "NV": "gt_NV",
                    "blink_id": "gt_blink_id",
                },
                inplace=True,
            )

        return df

    def generate_blink_ids(self, df: pd.DataFrame, blink_col: str, blink_id_col: str) -> pd.DataFrame:
        df[blink_id_col] = -1
        for eye in self.eyes:
            eye_str = eye.value.upper()
            in_blink = False
            blink_id_counter = 1
            for index, row in df[df["eye"] == eye_str].iterrows():
                is_blinking = row[blink_col] in [1, 2]

                if is_blinking and not in_blink:
                    in_blink = True
                    df.loc[index, blink_id_col] = blink_id_counter
                elif is_blinking and in_blink:
                    df.loc[index, blink_id_col] = blink_id_counter
                elif not is_blinking and in_blink:
                    in_blink = False
                    blink_id_counter += 1

        return df
